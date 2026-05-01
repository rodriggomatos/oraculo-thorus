"""Parser de XLSX — todas as abas convertidas pra markdown tables."""

import asyncio
from pathlib import Path

from openpyxl import load_workbook


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("|", "\\|").replace("\n", " ").strip()


def _sheet_to_markdown(sheet_name: str, rows: list[list[object]]) -> str:
    if not rows:
        return f"## {sheet_name}\n\n(aba vazia)"
    header = [_cell_to_str(c) for c in rows[0]]
    width = len(header)
    lines: list[str] = [f"## {sheet_name}", ""]
    lines.append("| " + " | ".join(header) + " |")
    lines.append("|" + "|".join(["---"] * width) + "|")
    for row in rows[1:]:
        padded = list(row) + [None] * (width - len(row))
        cells = [_cell_to_str(c) for c in padded[:width]]
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines)


def _extract_markdown_sync(file_path: Path) -> str:
    wb = load_workbook(filename=str(file_path), data_only=True, read_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        parts.append(_sheet_to_markdown(sheet_name, rows))
    wb.close()
    return "\n\n".join(parts)


async def parse(file_path: Path) -> str:
    return await asyncio.to_thread(_extract_markdown_sync, file_path)
